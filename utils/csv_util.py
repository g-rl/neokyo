import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

def export_clean_excel(csv_file, output_file="neokyo.xlsx"):
    df = pd.read_csv(csv_file)

    # real hyperlinks
    def make_hyperlink(url, text="LINK"):
        if pd.isna(url) or not isinstance(url, str) or not url.startswith("http"):
            return ""
        return f'=HYPERLINK("{url}", "{text}")'

    if "url" in df.columns:
        df["url"] = df["url"].apply(lambda x: make_hyperlink(x, "PRODUCT"))
    if "image_url" in df.columns:
        df["image_url"] = df["image_url"].apply(lambda x: make_hyperlink(x, "IMAGE"))

    df.to_excel(output_file, index=False, engine="openpyxl")
    wb = load_workbook(output_file)
    ws = wb.active

    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 25

    wb.save(output_file)
